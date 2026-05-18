"""XLSX parser: каждый лист → markdown-таблица.

Банковские тарифы часто в Excel. Игнорируем форматирование, берём только
значения. Каждый sheet — отдельная секция.
"""
from __future__ import annotations
import io, re
from openpyxl import load_workbook

from .base import ParsedDoc


_MAX_ROWS_PER_SHEET = 500
_MAX_SHEETS = 20


def _row_to_md(cells: list[str]) -> str:
    return "| " + " | ".join(c.strip().replace("\n", " ") for c in cells) + " |"


def parse_xlsx(content: bytes, url: str = "") -> ParsedDoc:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out_lines: list[str] = []
    sheet_names: list[str] = []

    for sheet in wb.sheetnames[:_MAX_SHEETS]:
        ws = wb[sheet]
        sheet_names.append(sheet)
        out_lines.append(f"\n## Лист: {sheet}\n")

        # Собираем значения построчно, отбрасывая полностью пустые строки
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= _MAX_ROWS_PER_SHEET:
                out_lines.append(f"\n_(обрезано: показаны первые {_MAX_ROWS_PER_SHEET} строк из {ws.max_row})_\n")
                break
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(cells)

        if not rows:
            continue

        # Нормализуем длину строк (чтобы markdown table работал)
        max_cols = max(len(r) for r in rows)
        for r in rows:
            while len(r) < max_cols:
                r.append("")

        # Первая непустая строка — header. Если в первой строке >50% числовых —
        # treat'им её как data, без header'а.
        header = rows[0]
        body_rows = rows[1:]
        numeric_in_header = sum(1 for c in header if re.fullmatch(r"[\d\s.,-]+", c.strip()))
        if numeric_in_header > len(header) // 2:
            header = [f"col{i+1}" for i in range(max_cols)]
            body_rows = rows

        out_lines.append(_row_to_md(header))
        out_lines.append("| " + " | ".join("---" for _ in header) + " |")
        for r in body_rows:
            out_lines.append(_row_to_md(r))

    body = "\n".join(out_lines)
    body = re.sub(r"\n{3,}", "\n\n", body).strip()

    return ParsedDoc(
        title=wb.properties.title if wb.properties else None,
        text=body, doc_type="xlsx",
        meta={"url": url, "sheets": sheet_names, "char_count": len(body)},
    )
