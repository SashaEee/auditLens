"""FastAPI APIRouter модуля loophole: эндпоинты + SSE-чат.

Префикс /api/loophole (монтируется в web/app.py). Авторизация — server-side:
trusted principal из X-Authentik-* (web/auth.py) + active membership и роли
из БД (loophole/authorization.py), перечитываемые на каждом запросе.
X-User-Id от клиента больше не доверяем.
"""
from __future__ import annotations

import logging
import os
import uuid
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sse_starlette.sse import EventSourceResponse

from .. import db
from ..web.auth import CurrentUser, get_current_user
from . import authorization, logging_audit
from . import collector as collector_mod
from . import refine as refine_mod
from . import repository as repo
from . import workspace as ws_mod
from .chat import clarify as clarify_mod
from .chat import graph as chat_graph
from .chat.state import ChatState
from .kb import repository as kb_repo
from .models import ExportRequest, SearchQuery, WorkspaceCreate
from .research_cases import ResearchCaseService

log = logging.getLogger(__name__)


# ── Dependencies ────────────────────────────────────────────────────────────
def get_session():
    """Yield SQLAlchemy-сессию. Переопределяется в тестах через
    app.dependency_overrides[get_session]."""
    with db.session() as s:
        yield s


def get_user_id(
    user: CurrentUser = Depends(get_current_user),
    session=Depends(get_session),
) -> str:
    """Единая граница авторизации модуля: trusted principal (X-Authentik-*
    от nginx) + active membership из БД. Возвращает username principal.
    401 — без аутентифицированного principal, 403 — без членства.
    Переопределяется в тестах через app.dependency_overrides."""
    principal = authorization.require_member(user, session=session)
    return principal.username


# Router-level guard: ВСЕ эндпоинты модуля требуют trusted principal +
# membership до чтения данных. Endpoint-level Depends(get_user_id) — тот же
# callable, FastAPI выполняет его один раз на запрос.
router = APIRouter(dependencies=[Depends(get_user_id)])


def _require_workspace_owner(workspace_id: int, user_id: str, *, session) -> None:
    """Ownership workspace: 404 — не существует, 403 — чужой."""
    ws = repo.get_workspace(workspace_id, session=session)
    if ws is None:
        raise HTTPException(status_code=404, detail="Рабочая область не найдена")
    if ws["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Нет доступа к чужому workspace")


# ── Рабочие контексты и очередь верификации (story 1.1) ─────────────────────
@router.get("/contexts")
def list_contexts(
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Доступные principal рабочие контексты: каталог и создание
    AI-исследования — любому члену, очередь — только эксперту ЦК КС."""
    return {"contexts": authorization.available_contexts(user_id, session=session)}


@router.get("/queue")
def verification_queue(
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Очередь верификации ЦК КС. Роль перечитывается из БД на каждый запрос:
    при отказе данные очереди не возвращаются."""
    authorization.require_role(
        user_id, authorization.ROLE_CCKS_EXPERT, action="queue_access", session=session,
    )
    records = repo.list_verification_queue(session=session)
    return {"records": records, "count": len(records)}


class SubmitResearchCandidateRequest(BaseModel):
    """Явный выбор evidence для неизменяемой передачи в ЦК КС."""

    evidence_source_ids: list[int] = Field(min_length=1)
    run_id: str = Field(min_length=1, max_length=200)


@router.post("/research/candidates/{candidate_id}/submit")
def submit_research_candidate(
    candidate_id: int,
    body: SubmitResearchCandidateRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Передаёт выбранную версию исследовательского кейса в очередь ЦК КС."""
    service = ResearchCaseService(session)
    workspace_id = service.candidate_workspace_id(candidate_id)
    if workspace_id is None:
        raise HTTPException(status_code=404, detail="Кандидат исследования не найден")
    _require_workspace_owner(workspace_id, user_id, session=session)
    snapshot = service.submit_for_verification(
        candidate_id,
        evidence_source_ids=body.evidence_source_ids,
        submitted_by=user_id,
        correlation_run_id=body.run_id,
    )
    if snapshot is None:
        raise HTTPException(
            status_code=409,
            detail="Выбранный источник недоступен. Вернитесь к разрешённым источникам.",
        )
    logging_audit.log_action(
        user_id,
        "submit_research_candidate",
        detail={"candidate_id": candidate_id, "snapshot_id": snapshot["snapshot_id"]},
        session=session,
    )
    return {**snapshot, "status_label": "Ожидает решения ЦК КС"}


@router.get("/research/reports/{report_id}/export/{format}")
async def export_research_report(
    report_id: int,
    format: str,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Скачивает PDF/DOCX только из server-side результата текущего workspace/run."""
    if format not in {"pdf", "docx"}:
        raise HTTPException(status_code=404, detail="Формат отчёта не поддерживается")
    report = ResearchCaseService(session).get_report_result(report_id)
    if report is None:
        raise HTTPException(status_code=404, detail="Отчёт исследования не найден")
    _require_workspace_owner(int(report["workspace_id"]), user_id, session=session)
    from . import pdf_export

    try:
        if format == "pdf":
            payload = await pdf_export.export_research_report_pdf(report)
            media_type = "application/pdf"
            filename = f"research-report-{report_id}.pdf"
        else:
            payload = pdf_export.export_research_report_docx(report)
            media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            filename = f"research-report-{report_id}.docx"
    except Exception as exc:
        log.warning("Экспорт отчёта исследования недоступен: %s", exc)
        code = "pdf_unavailable" if format == "pdf" else "docx_unavailable"
        message = "PDF-экспорт недоступен. Выберите Word или повторите PDF." if format == "pdf" else "Word-экспорт недоступен"
        raise HTTPException(status_code=503, detail={"code": code, "message": message}) from exc
    logging_audit.log_action(
        user_id,
        "research_report_export",
        workspace_id=int(report["workspace_id"]),
        detail={"report_id": report_id, "format": format},
        session=session,
    )
    return Response(
        content=payload,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class VerificationDecisionRequest(BaseModel):
    decision: str = Field(pattern="^(vulnerability|fraud_scheme|not_confirmed)$")
    comment: str = Field(min_length=1, max_length=4000)
    run_id: str = Field(min_length=1, max_length=200)


@router.post("/verification/snapshots/{snapshot_id}/decision")
def decide_verification_snapshot(
    snapshot_id: int,
    body: VerificationDecisionRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Фиксирует единственное решение эксперта для submitted snapshot."""
    authorization.require_role(
        user_id, authorization.ROLE_CCKS_EXPERT, action="verification_decide", session=session,
    )
    decision = ResearchCaseService(session).decide_snapshot(
        snapshot_id,
        decision=body.decision,
        comment=body.comment,
        decided_by=user_id,
        run_id=body.run_id,
    )
    if decision is None:
        raise HTTPException(status_code=409, detail="Итог по кейсу уже зафиксирован или недоступен")
    logging_audit.log_action(
        user_id,
        "verification_decision",
        detail={"snapshot_id": snapshot_id, "decision": decision["decision"]},
        session=session,
    )
    return decision


# ── Администрирование (story 1.5): роль ЦК КС, Telegram-цели, сводный аудит ──
def _require_admin(user_id: str, *, action: str, session) -> None:
    """Граница capability module_admin на КАЖДОМ админ-endpoint: 403 без
    активного назначения, отказ аудируется. Данные не возвращаются."""
    authorization.require_role(
        user_id, authorization.ROLE_MODULE_ADMIN, action=action, session=session,
        detail="Нет доступа к администрированию модуля",
    )


@router.get("/admin/roles")
def admin_roles(
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Управление ролью ЦК КС: назначения + счётчик активных экспертов."""
    _require_admin(user_id, action="admin_roles_read", session=session)
    return {
        "roles": authorization.list_ccks_assignments(session=session),
        "active_experts": authorization.count_active_ccks_experts(session=session),
        "max_experts": authorization.MAX_ACTIVE_CCKS_EXPERTS,
    }


class RoleChangeRequest(BaseModel):
    username: str = Field(min_length=1, max_length=150)

    @field_validator("username")
    @classmethod
    def normalize_username(cls, value: str) -> str:
        normalized = value.strip()
        if not normalized:
            raise ValueError("username не может быть пустым")
        return normalized

@router.post("/admin/roles/grant")
def admin_grant_role(
    body: RoleChangeRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Назначение роли ЦК КС. 409 — лимит пяти активных экспертов исчерпан.
    Изменение аудируется обезличенно (actor + действие + решение)."""
    _require_admin(user_id, action="role_grant", session=session)
    try:
        authorization.grant_ccks_expert(user_id, body.username, session=session)
    except authorization.ExpertLimitError as e:
        raise HTTPException(status_code=409, detail=str(e)) from e
    return {
        "username": body.username,
        "role": authorization.ROLE_CCKS_EXPERT,
        "status": "active",
        "active_experts": authorization.count_active_ccks_experts(session=session),
    }


@router.post("/admin/roles/revoke")
def admin_revoke_role(
    body: RoleChangeRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Отзыв роли ЦК КС: действует на следующий запрос очереди.
    404 — активного назначения нет. Изменение аудируется обезличенно."""
    _require_admin(user_id, action="role_revoke", session=session)
    revoked = authorization.revoke_ccks_expert(user_id, body.username, session=session)
    if not revoked:
        raise HTTPException(
            status_code=404, detail="Активное назначение роли ЦК КС не найдено",
        )
    return {
        "username": body.username,
        "role": authorization.ROLE_CCKS_EXPERT,
        "status": "revoked",
    }


@router.get("/admin/telegram-targets")
def admin_telegram_targets(
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Статус Telegram-целей: цель + операционный статус парсера,
    без технических payload."""
    _require_admin(user_id, action="admin_telegram_read", session=session)
    return {"targets": repo.list_telegram_targets(session=session)}


@router.get("/admin/audit")
def admin_audit(
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Сводный обезличенный аудит: агрегаты action/decision/count без
    username и payload. Само чтение аудита фиксируется в журнале."""
    _require_admin(user_id, action="admin_audit_read", session=session)
    authorization.log_auth_event(user_id, "admin_audit_read", "allow", session=session)
    return {"events": authorization.audit_summary(session=session)}


# ── Эндпоинты ───────────────────────────────────────────────────────────────
@router.post("/search")
def search(
    q: SearchQuery,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    records = repo.search_relevant(
        q.query_text,
        bank_slugs=q.bank_slugs or None,
        period_from=q.period_from,
        period_to=q.period_to,
        only_loophole=True,
        session=session,
    )
    logging_audit.log_action(
        user_id, "search",
        detail={"query": q.query_text, "banks": q.bank_slugs},
        session=session,
    )
    return {"records": records, "count": len(records)}


@router.get("/keywords")
def get_keywords(session=Depends(get_session)):
    return {"keywords": repo.list_keywords(session=session)}


@router.get("/records")
def list_records(
    bank_slugs: str | None = None,
    period_from: date | None = None,
    period_to: date | None = None,
    q: str | None = None,
    only_loophole: bool | None = None,
    status: str | None = None,
    limit: int = 500,
    offset: int = 0,
    session=Depends(get_session),
):
    """Список лазеек из БД для таблицы в основной области UI.

    bank_slugs передаётся строкой через запятую (query-param friendly):
    /records?bank_slugs=sberbank,vtb
    """
    slugs = (
        [s.strip() for s in bank_slugs.split(",") if s.strip()]
        if bank_slugs else None
    )
    records = repo.list_records(
        bank_slugs=slugs,
        period_from=period_from,
        period_to=period_to,
        query_text=q,
        only_loophole=only_loophole,
        status=status,
        limit=limit,
        offset=offset,
        session=session,
    )
    return {"records": records, "count": len(records)}


@router.get("/catalog")
def list_published_catalog(
    bank_slugs: str | None = None,
    period_from: date | None = None,
    period_to: date | None = None,
    q: str | None = None,
    limit: int = 500,
    offset: int = 0,
    session=Depends(get_session),
):
    """Общий каталог: только опубликованные подтверждённые кейсы."""
    slugs = [item.strip() for item in bank_slugs.split(",") if item.strip()] if bank_slugs else None
    records = repo.list_records(
        bank_slugs=slugs,
        period_from=period_from,
        period_to=period_to,
        query_text=q,
        only_loophole=True,
        status="published",
        limit=limit,
        offset=offset,
        session=session,
    )
    return {"records": records, "count": len(records)}


@router.get("/records/{record_id}/content")
def record_content(
    record_id: int,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Полный контент записи — ленивая подгрузка для разворачивающейся строки UI.

    raw_text в списках не отдаётся (payload); только здесь, по явному запросу.
    """
    record = repo.get_record(record_id, session=session)
    if record is None:
        raise HTTPException(status_code=404, detail="record not found")
    logging_audit.log_action(
        user_id, "view_content",
        detail={"record_id": record_id}, session=session,
    )
    return {
        "record_id": record_id,
        "content_status": record.get("content_status"),
        "raw_text": record.get("raw_text"),
        "raw_text_len": record.get("raw_text_len"),
        "raw_text_truncated": bool(record.get("raw_text_truncated")),
        "fetched_at": record.get("fetched_at"),
    }


class BackfillRequest(BaseModel):
    limit: int = 100
    delay_ms: int = 500


@router.post("/records/backfill-content")
def backfill_content(
    body: BackfillRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Догрузка полного контента для legacy/fetch_failed/empty записей.

    Синхронно, порциями limit с паузой delay_ms между fetch'ами (вежливый
    rate-limit). remaining в ответе — повторный вызов добирает хвост.
    fetch_failed остаётся в очереди; full/truncated выпадают навсегда.
    """
    import time

    from . import content_fetch

    targets = repo.list_records_needing_content(limit=body.limit, session=session)
    updated = 0
    failed = 0
    for i, row in enumerate(targets):
        if i:
            time.sleep(body.delay_ms / 1000)
        content = content_fetch.fetch_full_content(row["url"])
        if content.status == content_fetch.STATUS_FAILED:
            failed += 1
        else:
            updated += 1
        repo.update_content(
            row["record_id"],
            raw_text=content.text,
            content_status=content.status,
            raw_text_len=content.length,
            truncated=content.truncated,
            session=session,
        )
    remaining = repo.count_records_needing_content(session=session)
    logging_audit.log_action(
        user_id, "backfill_content",
        detail={"processed": len(targets), "updated": updated,
                "failed": failed, "remaining": remaining},
        session=session,
    )
    return {"processed": len(targets), "updated": updated,
            "failed": failed, "remaining": remaining}


class VerdictRequest(BaseModel):
    record_ids: list[int]
    is_loophole: bool
    comment: str | None = None


@router.post("/records/verdict")
def mark_verdict(
    body: VerdictRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Ручная маркировка записей: «лазейка» / «обычный запрос».

    Покрывает одиночную (массив из одного id) и массовую маркировку.
    is_loophole=true → пример добавляется в KB (дедуп по record_id);
    is_loophole=false → пример удаляется из KB (откат).
    """
    if not body.record_ids:
        raise HTTPException(status_code=400, detail="record_ids пуст")
    updated: list[int] = []
    skipped: list[int] = []
    reason = body.comment or f"manual:{user_id}"
    for rid in body.record_ids:
        record = repo.get_record(rid, session=session)
        if record is None:
            skipped.append(rid)
            continue
        repo.update_verdict(
            rid,
            is_loophole=body.is_loophole,
            confidence=1.0,
            reason=reason,
            model="manual",
            session=session,
        )
        if body.is_loophole:
            if repo.get_kb_example_by_record(rid, session=session) is None:
                description = (
                    record.get("snippet")
                    or (record.get("raw_text") or "")[:2000]
                    or record.get("title")
                    or ""
                )
                kb_repo.add_example(
                    record.get("title") or description[:200],
                    description,
                    category="manual",
                    record_id=rid,
                    session=session,
                )
        else:
            repo.delete_kb_example_by_record(rid, session=session)
        updated.append(rid)
    logging_audit.log_action(
        user_id, "mark_verdict",
        detail={
            "ids": body.record_ids,
            "is_loophole": body.is_loophole,
            "comment": body.comment,
        },
        session=session,
    )
    return {"updated": len(updated), "skipped": skipped}


@router.get("/banks")
def list_banks(session=Depends(get_session)):
    """Уникальные bank_slug из loophole_record — для фильтра таблицы."""
    return {"banks": repo.list_bank_slugs(session=session)}


@router.get("/workspaces")
def list_workspaces(user_id: str = Depends(get_user_id), session=Depends(get_session)):
    return {"workspaces": ws_mod.list_for_user(user_id, session=session)}


@router.post("/workspace")
def create_workspace(
    body: WorkspaceCreate,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    wid = ws_mod.create(user_id, name=body.name, session=session)
    logging_audit.log_action(
        user_id, "workspace_create", workspace_id=wid, session=session
    )
    return {"workspace_id": wid}


@router.get("/history/{workspace_id}")
def history(
    workspace_id: int,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    _require_workspace_owner(workspace_id, user_id, session=session)
    return {"messages": ws_mod.history(workspace_id, session=session)}


class ChatRequest(BaseModel):
    workspace_id: int
    message: str
    history: list[dict] = Field(default_factory=list)
    # true → уточнение уже пройдено (сообщение — обогащённый запрос после
    # /clarify/answer). Пропускаем clarify-гейт и идём выполнять. Без этого
    # /chat заново гонял бы generate_clarifications на КАЖДЫЙ вызов → петля.
    # Одноразовый server-side token, выданный после ответа на clarification.
    clarify_token: str | None = None


@router.post("/chat")
async def chat(
    body: ChatRequest,
    request: Request,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """SSE-чат: стримит token/tool_call/tool_result/record события."""
    _require_workspace_owner(body.workspace_id, user_id, session=session)
    clarification_verified = clarify_mod.consume_execution_token(
        body.clarify_token,
        user_id=user_id,
        workspace_id=body.workspace_id,
        query=body.message,
    )
    if body.clarify_token is not None and not clarification_verified:
        logging_audit.log_action(
            user_id,
            "chat_rejected",
            workspace_id=body.workspace_id,
            detail={"reason": "invalid_execution_token"},
            session=session,
        )
        raise HTTPException(status_code=400, detail="Недействительный execution token")
    state: ChatState = {
        "query": body.message,
        "messages": body.history,
        "workspace_id": body.workspace_id,
        "user_id": user_id,
        "session": session,
        "clarification_verified": clarification_verified,
        "run_id": uuid.uuid4().hex,
    }
    # Обогащённый execution input уже представлен в истории исходным запросом
    # и ответом на clarification, поэтому не сохраняем его третьим user-message.
    if not clarification_verified:
        repo.add_chat_message(body.workspace_id, "user", body.message, session=session)
    logging_audit.log_action(
        user_id, "chat", workspace_id=body.workspace_id,
        detail={"message": repo.redact_audit_text(body.message, limit=200)}, session=session,
    )

    async def event_generator():
        import json as _json
        report_chunks: list[str] = []
        try:
            stream = chat_graph.stream_chat(state, session=session)
            async for ev in stream:
                if ev["event"] in {"token", "partial"}:
                    data = ev["data"]
                    piece = data if isinstance(data, str) else data.get("text", data.get("message", ""))
                    if isinstance(piece, str):
                        report_chunks.append(piece)
                yield {
                    "event": ev["event"],
                    "data": _json.dumps(ev["data"], ensure_ascii=False, default=str),
                }
            result_text = state.get("answer") or "".join(report_chunks)
            if result_text and state.get("run_id"):
                report_id = ResearchCaseService(session).save_report_result(
                    workspace_id=body.workspace_id,
                    run_id=str(state["run_id"]),
                    query=body.message,
                    result=str(result_text),
                )
                yield {"event": "report", "data": _json.dumps({"report_id": report_id})}
            # Сохраняем ответ (если есть).
            try:
                if state.get("answer"):
                    repo.add_chat_message(
                        body.workspace_id, "assistant", state["answer"], session=session
                    )
            except Exception:
                pass
        finally:
            close_stream = getattr(locals().get("stream"), "aclose", None)
            if callable(close_stream):
                try:
                    await close_stream()
                except Exception:
                    log.warning("[chat] закрытие graph stream завершилось ошибкой")

    return EventSourceResponse(event_generator())


EXPORT_LIMIT = 10000  # максимум записей в одной выгрузке
# Сколько знаков текста страницы кладём в CSV на одну запись.
_CSV_TEXT_LIMIT = int(os.getenv("LOOPHOLE_CSV_TEXT_LIMIT", "10000"))


def _csv_safe_cell(value: object) -> object:
    """Закрывает строковую ячейку от интерпретации как формулы Excel."""
    if isinstance(value, str) and value.lstrip(" \t").startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


@router.post("/export")
def export(
    body: ExportRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    if body.records and len(body.records) > EXPORT_LIMIT:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Нельзя выгрузить более {EXPORT_LIMIT} записей за раз "
                f"(запрошено {len(body.records)}). Сократите выделение."
            ),
        )
    records = []
    if body.records:
        for rid in body.records:
            r = repo.get_record(rid, session=session)
            if r and r.get("status") == "published" and r.get("is_loophole") is True:
                records.append(r)
    logging_audit.log_action(
        user_id, "export", detail={"format": body.format, "count": len(records)},
        session=session,
    )
    if body.format == "json":
        return JSONResponse(records)
    if body.format == "csv":
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow([
            "record_id", "title", "url", "domain", "bank_slug", "keyword",
            "is_loophole", "verdict_confidence",
            "verdict_reason", "verdict_model", "status",
            "published_at", "collected_at", "classified_at",
            "content_status", "raw_text_len", "raw_text",
        ])
        for r in records:
            row = [
                r.get("record_id"), r.get("title"), r.get("url"),
                r.get("domain"), r.get("bank_slug"), r.get("keyword"),
                r.get("is_loophole"), r.get("verdict_confidence"), r.get("verdict_reason"),
                r.get("verdict_model"), r.get("status"),
                r.get("published_at"), r.get("collected_at"), r.get("classified_at"),
                r.get("content_status"), r.get("raw_text_len"),
                # текст страницы обрезаем: полный (до 200 000 знаков на запись)
                # ×  тысячи записей строится в памяти единственного процесса —
                # выгрузка ронял бы весь AuditLens, а не только «Лазейки».
                (r.get("raw_text") or "")[:_CSV_TEXT_LIMIT],
            ]
            writer.writerow([_csv_safe_cell(value) for value in row])
        # BOM для корректного открытия в Excel (Windows).
        return Response(
            content="\ufeff" + buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=loopholes.csv"},
        )
    # pdf — через pdf_export (Playwright); заглушка для тестов.
    return JSONResponse({"error": "pdf export requires Playwright"}, status_code=501)


class ReportFilterV1(BaseModel):
    """Единый фильтр опубликованного каталога и его выгрузок."""

    bank_slugs: list[str] = Field(default_factory=list)
    period_from: date | None = None
    period_to: date | None = None
    query_text: str = ""
    only_loophole: bool | None = None
    status: str | None = None


# Обратная совместимость прежнего имени request-модели.
FilteredExportRequest = ReportFilterV1


class AnalyticsQueryRequest(BaseModel):
    sql: str = Field(min_length=1, max_length=2000)
    params: dict[str, object] = Field(default_factory=dict)


class ScheduledAnalyticsRequest(BaseModel):
    """Запрос содержит только имя серверной задачи, но не raw SQL."""

    model_config = ConfigDict(extra="forbid")

    query_id: str = Field(min_length=1, max_length=100)
    workspace_id: int = Field(gt=0)
    recipient_username: str = Field(min_length=1, max_length=200)
    cron_expr: str = Field(min_length=1, max_length=100)
    expires_at: datetime


@router.post("/export/csv")
def export_csv_filtered(
    body: ReportFilterV1,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Выгрузка CSV по текущим фильтрам таблицы (без передачи ids).
    Берёт все подходящие записи (limit 10000) и формирует CSV с BOM."""
    records = repo.list_records(
        bank_slugs=body.bank_slugs or None,
        period_from=body.period_from,
        period_to=body.period_to,
        query_text=body.query_text or None,
        only_loophole=True,
        status="published",
        limit=10000,
        include_content=True,
        session=session,
    )
    logging_audit.log_action(
        user_id, "export_csv", detail={"count": len(records)}, session=session,
    )
    import csv as _csv
    import io as _io
    buf = _io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow([
        "record_id", "title", "url", "domain", "bank_slug", "keyword",
        "trust_score", "is_loophole", "verdict_confidence",
        "verdict_reason", "verdict_model", "status",
        "collected_at", "classified_at",
        "content_status", "raw_text_len", "raw_text",
    ])
    for r in records:
        writer.writerow([
            r.get("record_id"), r.get("title"), r.get("url"),
            r.get("domain"), r.get("bank_slug"), r.get("keyword"),
            r.get("trust_score"), r.get("is_loophole"),
            r.get("verdict_confidence"), r.get("verdict_reason"),
            r.get("verdict_model"), r.get("status"),
            r.get("collected_at"), r.get("classified_at"),
            r.get("content_status"), r.get("raw_text_len"), r.get("raw_text"),
        ])
    return Response(
        content="\ufeff" + buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=loopholes.csv"},
    )


@router.post("/export/xlsx")
def export_xlsx_filtered(
    body: ReportFilterV1,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Выгружает published-каталог в XLSX без неполных файлов."""
    records = repo.list_records(
        bank_slugs=body.bank_slugs or None,
        period_from=body.period_from,
        period_to=body.period_to,
        query_text=body.query_text or None,
        only_loophole=True,
        status="published",
        limit=10_001,
        include_content=True,
        session=session,
    )
    if len(records) > 10_000:
        raise HTTPException(
            status_code=409,
            detail=f"Найдено {len(records)} записей. Сузьте фильтры до 10000 для XLSX.",
        )
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Опубликованные кейсы")
    columns = ("record_id", "title", "url", "bank_slug", "snippet", "status")
    sheet.append(list(columns))
    for record in records:
        sheet.append([record.get(column) for column in columns])
    buffer = BytesIO()
    workbook.save(buffer)
    logging_audit.log_action(user_id, "export_xlsx", detail={"count": len(records)}, session=session)
    return Response(
        content=buffer.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": "attachment; filename=loopholes.xlsx"},
    )


@router.post("/export/pdf")
async def export_pdf_filtered(
    body: ReportFilterV1,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Рендерит PDF того же published-каталога, что таблица и CSV/XLSX."""
    records = repo.list_records(
        bank_slugs=body.bank_slugs or None,
        period_from=body.period_from,
        period_to=body.period_to,
        query_text=body.query_text or None,
        only_loophole=True,
        status="published",
        limit=10_000,
        include_content=False,
        session=session,
    )
    from .pdf_export import export_pdf

    try:
        payload = await export_pdf(records)
    except Exception as exc:
        log.warning("PDF-экспорт недоступен: %s", exc)
        raise HTTPException(
            status_code=503,
            detail={"code": "pdf_unavailable", "message": "PDF-экспорт недоступен"},
        ) from exc
    logging_audit.log_action(user_id, "export_pdf", detail={"count": len(records)}, session=session)
    return Response(
        content=payload,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=loopholes.pdf"},
    )


@router.post("/analytics/query")
def analytics_query(
    body: AnalyticsQueryRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Выполняет allowlisted параметризованный SELECT published view."""
    from .analytics import AnalyticsQueryError, execute_analytics_query

    try:
        result = execute_analytics_query(body.sql, body.params, session=session)
    except AnalyticsQueryError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    logging_audit.log_action(user_id, "analytics_query", detail={"rows": len(result["rows"])}, session=session)
    return result


@router.get("/analytics/tasks")
def list_analytics_tasks():
    """Именованные задачи, которые можно включить по расписанию."""
    from .scheduled_analytics import available_named_queries

    return {"tasks": available_named_queries()}


@router.post("/analytics/schedules")
def enable_analytics_schedule(
    body: ScheduledAnalyticsRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Создаёт версионированный внутренний ScheduledQueryContract v1."""
    from .scheduled_analytics import ScheduledAnalyticsService

    _require_workspace_owner(body.workspace_id, user_id, session=session)
    if not authorization.is_active_member(body.recipient_username, session=session):
        raise HTTPException(status_code=422, detail="Получатель не является активным членом модуля")
    try:
        contract = ScheduledAnalyticsService(session).enable(
            query_id=body.query_id,
            workspace_id=body.workspace_id,
            owner_username=user_id,
            recipient_username=body.recipient_username,
            cron_expr=body.cron_expr,
            expires_at=body.expires_at,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    logging_audit.log_action(
        user_id,
        "analytics_schedule_enabled",
        workspace_id=body.workspace_id,
        detail={"scheduled_query_id": contract["scheduled_query_id"], "query_id": body.query_id},
        session=session,
    )
    return contract


@router.get("/analytics/schedules/{scheduled_query_id}/results")
def scheduled_analytics_results(
    scheduled_query_id: int,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Чтение внутреннего результата только владельцем либо указанным получателем."""
    from .scheduled_analytics import ScheduledAnalyticsService

    try:
        results = ScheduledAnalyticsService(session).list_results(
            scheduled_query_id,
            username=user_id,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    return {"results": results}


@router.post("/refine")
async def refine(
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    added = await refine_mod.refine_keywords(session=session)
    logging_audit.log_action(
        user_id, "refine", detail={"added": added}, session=session
    )
    return {"added": added}


@router.post("/collect/run")
async def collect_run(
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Ручной запуск авто-сборщика (для админа)."""
    n = await collector_mod.collect_once(session=session)
    logging_audit.log_action(
        user_id, "collect", detail={"new_records": n}, session=session
    )
    return {"new_records": n}


# ── Clarify-воронка ─────────────────────────────────────────────────────────
class ClarifyRequest(BaseModel):
    question: str
    history: list[dict] = Field(default_factory=list)
    workspace_id: int | None = None


class ClarifyAnswerRequest(BaseModel):
    workspace_id: int
    question: str
    answers: list[dict] = Field(default_factory=list)
    clarification_token: str


@router.post("/clarify")
async def clarify(
    body: ClarifyRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Генерация уточняющих вопросов по запросу аудитора."""
    if body.workspace_id is not None:
        _require_workspace_owner(body.workspace_id, user_id, session=session)

    result = await clarify_mod.generate_clarifications(
        body.question, history=body.history
    )
    questions = clarify_mod.clarification_questions(result, query=body.question)
    if questions:
        result = {**result, "questions": questions}
    if (
        body.workspace_id is not None
        and not result.get("complete")
        and questions
    ):
        result = {
            **result,
            "clarification_token": clarify_mod.issue_clarification_token(
                user_id=user_id,
                workspace_id=body.workspace_id,
                query=body.question,
            ),
        }
    logging_audit.log_action(
        user_id, "clarify",
        detail={
            "question": repo.redact_audit_text(body.question, limit=200),
            "complete": result.get("complete"),
        },
        workspace_id=body.workspace_id,
        session=session,
    )
    return result


@router.post("/clarify/answer")
async def clarify_answer(
    body: ClarifyAnswerRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Сборка обогащённого запроса из исходного вопроса и ответов воронки."""
    _require_workspace_owner(body.workspace_id, user_id, session=session)
    answered = clarify_mod._answers_summary(body.answers)
    if not answered:
        logging_audit.log_action(
            user_id,
            "clarify_answer",
            workspace_id=body.workspace_id,
            detail={
                "question": repo.redact_audit_text(body.question, limit=200),
                "complete": False,
                "reason": "answers_required",
            },
            session=session,
        )
        return clarify_mod._clarification_answers_required()
    if not clarify_mod.validate_clarification_token(
        body.clarification_token,
        user_id=user_id,
        workspace_id=body.workspace_id,
        query=body.question,
    ):
        raise HTTPException(
            status_code=400,
            detail="Уточнение устарело или не принадлежит этому запросу",
        )

    try:
        enriched = await clarify_mod.build_enriched_question(body.question, body.answers)
    except Exception:
        log.warning("[clarify_answer] deterministic assembly failed — fail-closed")
        logging_audit.log_action(
            user_id,
            "clarify_answer",
            workspace_id=body.workspace_id,
            detail={
                "question": repo.redact_audit_text(body.question, limit=200),
                "complete": False,
                "reason": "clarification_assembly_failed",
            },
            session=session,
        )
        raise HTTPException(
            status_code=503,
            detail={
                "code": "clarification_assembly_failed",
                "message": "Не удалось подготовить исследование. Повторите отправку ответа.",
            },
        ) from None
    if isinstance(enriched, dict):
        logging_audit.log_action(
            user_id,
            "clarify_answer",
            workspace_id=body.workspace_id,
            detail={
                "question": repo.redact_audit_text(body.question, limit=200),
                "complete": False,
                "reason": enriched.get("reason", "clarification_unavailable"),
            },
            session=session,
        )
        raise HTTPException(
            status_code=503,
            detail={
                "code": "clarification_assembly_failed",
                "message": "Не удалось подготовить исследование. Повторите отправку ответа.",
            },
        )
    # Поглощаем challenge только после успешной детерминированной сборки:
    # внутренний 503 оставляет тот же ownership-bound token пригодным для retry.
    if not clarify_mod.consume_clarification_token(
        body.clarification_token,
        user_id=user_id,
        workspace_id=body.workspace_id,
        query=body.question,
    ):
        raise HTTPException(
            status_code=400,
            detail="Уточнение устарело или не принадлежит этому запросу",
        )
    answer_message = "; ".join(
        ", ".join(answer["vals"])
        for answer in answered
    )
    repo.add_chat_message(
        body.workspace_id,
        "user",
        answer_message,
        session=session,
    )
    execution_token = clarify_mod.issue_execution_token(
        user_id=user_id,
        workspace_id=body.workspace_id,
        query=enriched,
    )
    logging_audit.log_action(
        user_id, "clarify_answer",
        workspace_id=body.workspace_id,
        detail={
            "question": repo.redact_audit_text(body.question, limit=200),
            "enriched_len": len(enriched),
        },
        session=session,
    )
    return {
        "enriched_question": enriched,
        "execution_token": execution_token,
        "answer_message": answer_message,
    }


# ── Парсеры: общий каталог ──────────────────────────────────────────────────
class ParserCreateRequest(BaseModel):
    workspace_id: int
    query: str


class ParserPatchRequest(BaseModel):
    name: str | None = None
    cron_expr: str | None = None
    auto_enabled: bool | None = None


@router.get("/parsers")
def list_parsers(session=Depends(get_session)):
    """Общий каталог парсеров всех пользователей + статистика карточек."""
    from .parsers import registry as parser_registry

    return {"parsers": parser_registry.list_catalog(session=session)}


@router.post("/parsers")
async def create_parser(
    body: ParserCreateRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Создание парсера: дедуп источников (409 при полном дубле) + LLM-генерация."""
    from .parsers import dedup as dedup_mod
    from .parsers import generator as parser_generator
    from .parsers import registry as parser_registry

    targets = parser_generator.extract_targets(body.query)
    if not targets:
        raise HTTPException(
            status_code=422,
            detail="В запросе не указан URL ресурса или группа мессенджера",
        )
    keys = [k for k in (dedup_mod.normalize_target(t) for t in targets) if k]
    conflicts = parser_registry.find_conflicts(keys, session=session)
    full = [c for c in conflicts if sorted(c["overlap"]) == sorted(keys)]
    if full:
        c = full[0]
        raise HTTPException(
            status_code=409,
            detail={
                "error": "duplicate",
                "conflict_with": {"parser_id": c["parser_id"], "name": c["name"]},
            },
        )
    try:
        # session=None: фоновая валидация переживает запрос; repo-функции сами
        # открывают db.session() с commit (request-session не коммитит фон).
        result = await parser_generator.generate_parser(
            user_id, body.workspace_id, body.query, session=None
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e
    if conflicts:
        result["warnings"] = [
            {"target": o, "conflict_with": c["parser_id"]}
            for c in conflicts for o in c["overlap"]
        ]
    logging_audit.log_action(
        user_id, "parser_create",
        workspace_id=body.workspace_id,
        detail={"parser_id": result.get("parser_id"), "query": body.query[:200]},
        session=session,
    )
    return {
        "parser_id": result["parser_id"],
        "validation_run_id": result["validation_run_id"],
        "name": result["name"],
        "targets": result["targets"],
        "warnings": result.get("warnings") or [],
    }


@router.patch("/parsers/{parser_id}")
def patch_parser(
    parser_id: int,
    body: ParserPatchRequest,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Редактирование: имя, cron (валидация croniter), вкл/выкл автозапуска."""
    from .parsers import scheduler as parser_scheduler

    row = repo.get_parser(parser_id, session=session)
    if row is None:
        raise HTTPException(status_code=404, detail="parser not found")
    # cron_expr: None в теле = не менять; "" = очистить (расписание не настроено).
    if body.cron_expr is not None:
        cron_expr = body.cron_expr or None
    else:
        cron_expr = row.get("cron_expr")
    auto_enabled = (
        body.auto_enabled
        if body.auto_enabled is not None
        else bool(row.get("auto_enabled"))
    )
    if auto_enabled and row.get("status") != "ready":
        raise HTTPException(
            status_code=409,
            detail="Расписание доступно только после успешной валидации парсера",
        )
    nxt = None
    if cron_expr:
        try:
            nxt = parser_scheduler.next_run(cron_expr)
        except ValueError as e:
            raise HTTPException(status_code=422, detail=str(e)) from e
    if not auto_enabled:
        nxt = None
    repo.update_parser_schedule(
        parser_id,
        cron_expr=cron_expr,
        auto_enabled=auto_enabled,
        next_run_at=nxt,
        last_edited_by=user_id,
        name=body.name,
        session=session,
    )
    logging_audit.log_action(
        user_id, "parser_edit",
        detail={"parser_id": parser_id, "cron": cron_expr, "auto": auto_enabled},
        session=session,
    )
    updated = repo.get_parser(parser_id, session=session)
    updated["next_run_at"] = nxt.isoformat() if nxt else None
    return {"parser": updated}


@router.post("/parsers/{parser_id}/run")
async def run_parser(
    parser_id: int,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Ручной запуск парсера. Возвращает run_id для SSE-подписки на лог."""
    from .parsers import runner as runner_mod

    parser = repo.get_parser(parser_id, session=session)
    if parser is None:
        raise HTTPException(status_code=404, detail="parser not found")
    if parser.get("status") != "ready":
        raise HTTPException(
            status_code=409,
            detail="Парсер не прошёл успешную валидацию и не может быть запущен",
        )
    try:
        # session=None: фоновая wait() переживает запрос; repo-функции сами
        # открывают db.session() с commit (request-session не коммитит фон).
        run_id = await runner_mod.run(parser_id, "manual")
    except RuntimeError as e:
        raise HTTPException(status_code=409, detail=str(e)) from e
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e)) from e
    logging_audit.log_action(
        user_id, "parser_run",
        detail={"parser_id": parser_id, "run_id": run_id},
        session=session,
    )
    return {"parser_id": parser_id, "run_id": run_id}


@router.post("/parsers/{parser_id}/stop")
async def stop_parser(
    parser_id: int,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Останов запущенного парсера. 404 если не running."""
    from .parsers.runner import _RUNNING

    runner = _RUNNING.get(parser_id)
    if runner is None:
        raise HTTPException(status_code=404, detail="parser not running")
    await runner.stop()
    logging_audit.log_action(
        user_id, "parser_stop", detail={"parser_id": parser_id}, session=session,
    )
    return {"parser_id": parser_id, "stopped": True}


@router.get("/parsers/{parser_id}/status")
async def parser_status(
    parser_id: int,
    session=Depends(get_session),
):
    """Статус парсера: runtime (если running) + запись из БД."""
    from .parsers import registry as parser_registry
    from .parsers.runner import _RUNNING

    runner = _RUNNING.get(parser_id)
    if runner is not None:
        runtime = await runner.status()
    else:
        runtime = None
    row = parser_registry.get_parser(parser_id, session=session)
    if row is None and runtime is None:
        raise HTTPException(status_code=404, detail="parser not found")
    return {"parser_id": parser_id, "runtime": runtime, "parser": row}


@router.get("/parsers/{parser_id}/runs")
def list_parser_runs(parser_id: int, session=Depends(get_session)):
    """История запусков (последние 20)."""
    return {"runs": repo.list_runs(parser_id, session=session)}


@router.get("/parsers/{parser_id}/log/stream")
async def parser_log_stream(parser_id: int, run_id: int):
    """SSE-стрим лога запуска: события 'log' (строка) и 'done' (финал)."""
    from .parsers import runner as runner_mod

    queue = runner_mod.subscribe(run_id)

    async def events():
        try:
            for line in runner_mod.log_tail(run_id):
                yield {"event": "log", "data": line}
            while True:
                msg = await queue.get()
                yield msg
                if msg.get("event") == "done":
                    break
        finally:
            runner_mod.unsubscribe(run_id, queue)

    return EventSourceResponse(events())


@router.post("/parsers/{parser_id}/heal")
async def heal_parser(
    parser_id: int,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Ручной запуск анализа и восстановления парсера nanobot'ом."""
    from .parsers import healer as healer_mod

    if not healer_mod.nanobot_available():
        raise HTTPException(status_code=503, detail="nanobot unavailable")
    try:
        run_id = await healer_mod.heal(parser_id, manual=True, session=session)
    except RuntimeError as e:
        raise HTTPException(status_code=409, detail=str(e)) from e
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e)) from e
    logging_audit.log_action(
        user_id, "parser_heal",
        detail={"parser_id": parser_id, "run_id": run_id},
        session=session,
    )
    return {"parser_id": parser_id, "heal_run_id": run_id}


@router.delete("/parsers/{parser_id}")
def delete_parser(
    parser_id: int,
    user_id: str = Depends(get_user_id),
    session=Depends(get_session),
):
    """Удаление парсера (код + запись БД). 409 если running, 404 если не найден."""
    from .parsers import registry as parser_registry
    from .parsers.runner import _RUNNING

    if parser_id in _RUNNING:
        raise HTTPException(status_code=409, detail="parser running")
    deleted = parser_registry.delete_parser(parser_id, session=session)
    if not deleted:
        raise HTTPException(status_code=404, detail="parser not found")
    logging_audit.log_action(
        user_id, "parser_delete", detail={"parser_id": parser_id}, session=session,
    )
    return {"deleted": True}


# ── Загрузка таблицы по фильтрам (для агента) ───────────────────────────────
class TableLoadRequest(BaseModel):
    bank_slugs: list[str] = Field(default_factory=list)
    period_from: date | None = None
    period_to: date | None = None
    query_text: str = ""
    only_loophole: bool | None = None
    status: str | None = None
    limit: int = 500
    offset: int = 0


@router.post("/table/load")
def table_load(body: TableLoadRequest, session=Depends(get_session)):
    """Применяет фильтры и возвращает records для таблицы UI."""
    records = repo.list_records(
        bank_slugs=body.bank_slugs or None,
        period_from=body.period_from,
        period_to=body.period_to,
        query_text=body.query_text or None,
        only_loophole=body.only_loophole,
        status=body.status,
        limit=body.limit,
        offset=body.offset,
        session=session,
    )
    return {"records": records, "count": len(records)}
